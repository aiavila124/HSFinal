#Explorar datos
import pandas as pd
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from sklearn.ensemble import IsolationForest


#Clasificacion
from sklearn.model_selection import train_test_split
from sklearn.preprocessing import OneHotEncoder
from sklearn.preprocessing import StandardScaler
from sklearn.linear_model import LogisticRegression
import random

#Para imprimir resultados
import json
import os

class Modelo():
    def __init__(self,app):
        self.active = 1
    
    def codigo_modelo(self,file):
        file.save('temp/' + file.filename)
        # dataFile.save('temp/' + dataFile.filename)
        # print("Aqui va el valor del file", file.filename)
        routeFinal = os.path.join('temp', 'prueba.xlsx')
        df1= pd.read_excel(routeFinal,index_col=0)
        
        num=[ #Financieras
            'costo_cobro', 'num_acuerdos_cumplidos', 'abono_total_deuda','Capital_vencido_actual',
            'edad_mora_actual', 'Meta_recaudo_max','capital_max', 'valor_cuota_max',  'Cartera leve', 'Cartera regular', 'Cartera grave',
            #Gestiones
            'num_promesa_pago', 'num_ya pago', 'inconvenientes', 'num_titular_contacto','num_gestiones_contacto', 
            'num_llamada_contacto', 'num_visita_contacto','num_wapp_contacto']

        cat =[#Financieras
            'Alivio', 'Vencido_hoy','Cartera especial', 
            #Gestiones
            'moda_estado_cliente','medio_moda', 'moda_gestion', 
            'moda_tipo_contacto', 'medio_promesa']
        
        df1[cat] = df1[cat].astype('category')
        df1['sub_Cluster'] = df1['sub_Cluster'].astype('category')

        X = df1[['costo_cobro', 'num_acuerdos_cumplidos', 'abono_total_deuda',
            'Cartera leve', 'Cartera regular', 'Cartera grave', 'Cartera especial',
            'Alivio','Vencido_hoy','Capital_vencido_actual', 'edad_mora_actual',
            'Meta_recaudo_max', 'capital_max','valor_cuota_max', 'moda_estado_cliente', 
            'medio_moda', 'moda_gestion', 'moda_tipo_contacto', 'medio_promesa', 'num_promesa_pago',
            'num_ya pago', 'num_titular_contacto', 'num_gestiones_contacto',
            'num_llamada_contacto', 'num_visita_contacto', 'num_wapp_contacto',
            'inconvenientes']]
        y = df1['sub_Cluster']

        #Training-Testing estratificado
        X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.3, stratify=y)

        print(np.bincount(y_train))

        #Hacer numericas las categoricas

        # Crear el objeto OneHotEncoder
        onehot = OneHotEncoder(handle_unknown='ignore')

        # Concatenar los datos de entrenamiento y prueba para ajustar el encoder
        data = pd.concat([X_train[cat], X_test[cat]], axis=0)

        # Ajustar el encoder a los datos concatenados
        onehot.fit(data)
        new_columns = onehot.get_feature_names_out(cat)

        # Aplicar el encoder a los datos de entrenamiento y prueba
        X_train_enc = onehot.transform(X_train[cat])
        X_test_enc = onehot.transform(X_test[cat])

        #Crear df con el nombre de las columnas
        X_train_cat = pd.DataFrame(X_train_enc.toarray(), columns=new_columns)
        X_test_cat = pd.DataFrame(X_test_enc.toarray(), columns=new_columns)


        #Normalizar variables
        scaler = StandardScaler()
        X_train_scaled = scaler.fit_transform(X_train[num])
        X_test_scaled= scaler.transform(X_test[num])
        X_train_num = pd.DataFrame(X_train_scaled, columns=num)
        X_test_num = pd.DataFrame(X_test_scaled, columns=num)

        #Unir ambas normalizaciones
        X_train_df= pd.concat([X_train_cat, X_train_num], axis=1)
        X_test_df= pd.concat([X_test_cat, X_test_num], axis=1)

        """Modelo de clasificación con regresión logistica multinominal"""
        # Definir la semilla global
        random.seed(1234)
        y_test_df=pd.DataFrame(y_test)

        # Definir el modelo de regresión logística multinomial
        model_log= LogisticRegression(multi_class='multinomial', solver='lbfgs') #lbfgs, newton-cg, liblinear, sag
        model_log.fit(X_train_df, y_train)

        # Entrenar el modelo
        model_log.fit(X_train_df, y_train)

        # Predicción para comprobar
        y_pred_log = model_log.predict(X_test_df)
        y_pred_prob_log = model_log.predict_proba(X_test_df.values)
        y_prob_pos = y_pred_prob_log[:, 1]  

        #Tabla de cluster resultantes de reg_log
        y_test_df=pd.DataFrame(y_test)
        y_test_df1=pd.DataFrame(y_test)
        #print(y_pred.shape) 

        #Obtener los nombres de los clientes del conjunto de datos original:
        nombres_clientes = df1.loc[y_test_df.index, 'cod_neg']

        # Crear un DataFrame con los clusters predichos y los originales
        clientes_clusters_df = pd.DataFrame({'id': nombres_clientes.values})
        clientes_clusters_df['cluster_calculado_red'] = y_pred_log
        clientes_clusters_df['Cluster original'] = y_test_df1.reset_index(drop=True)
        clientes_clusters_df

        
        X_nuevos = pd.read_excel('temp/' + file.filename)

        X_nuevos_enc = onehot.transform(X_nuevos[cat])
        X_nuevos_cat = pd.DataFrame(X_nuevos_enc.toarray(), columns=new_columns)

        #Normalizar variables
        scaler = StandardScaler()
        scaler.fit(X_train[num])
        X_nuevos_scaled= scaler.transform(X_nuevos[num])
        
        X_nuevos_num = pd.DataFrame(X_nuevos_scaled, columns=num)

        #Unir ambas normalizaciones
        X_nuevos_df= pd.concat([X_nuevos_cat, X_nuevos_num], axis=1)

        #Herramienta para Predecir a que grupo pertenecen los clientes del df nuevo
        predicciones = model_log.predict(X_nuevos_df)

        #Obtener los nombres de los clientes del conjunto de datos original:
        id_clientes = df1.loc[X_nuevos.index, 'cod_neg']

        #Crear un DataFrame con los clusters predichos y los originales
        clientes_clusters = pd.DataFrame({'id': id_clientes.values})
        clientes_clusters['cluster'] = predicciones


        cluster1 = clientes_clusters.loc[clientes_clusters['cluster'] == 1, ['id']].rename(columns={'id': 'Código Negocio'})
        cluster2 = clientes_clusters.loc[clientes_clusters['cluster'] == 2, ['id']].rename(columns={'id': 'Código Negocio'})
        cluster3 = clientes_clusters.loc[clientes_clusters['cluster'] == 3, ['id']].rename(columns={'id': 'Código Negocio'})
        cluster4 = clientes_clusters.loc[clientes_clusters['cluster'] == 4, ['id']].rename(columns={'id': 'Código Negocio'})
        cluster5 = clientes_clusters.loc[clientes_clusters['cluster'] == 5, ['id']].rename(columns={'id': 'Código Negocio'})
        cluster6 = clientes_clusters.loc[clientes_clusters['cluster'] == 6, ['id']].rename(columns={'id': 'Código Negocio'})

        workbook = Workbook()
        hoja1 = workbook.active
        hoja1.title = "Excelentes"
        hoja2 = workbook.create_sheet(title="Buenos")
        hoja3 = workbook.create_sheet(title="Recuperables Buenos")
        hoja4 = workbook.create_sheet(title="Recuperables Graves")
        hoja5 = workbook.create_sheet(title="Graves")
        hoja6 = workbook.create_sheet(title="Muy difíciles")

        for row in dataframe_to_rows(cluster1, index=False, header=True):
            hoja1.append(row)

        for row in dataframe_to_rows(cluster2, index=False, header=True):
            hoja2.append(row)

        for row in dataframe_to_rows(cluster3, index=False, header=True):
            hoja3.append(row)
        
        for row in dataframe_to_rows(cluster4, index=False, header=True):
            hoja4.append(row)

        for row in dataframe_to_rows(cluster5, index=False, header=True):
            hoja5.append(row)
        
        for row in dataframe_to_rows(cluster6, index=False, header=True):
            hoja6.append(row)      

        ruta = os.path.join('temp', 'Lista_Clientes.xlsx')
        print('holaaaaa',ruta)
        workbook.save(ruta)

        clusters = {}
        for cluster in range(6):
            ids_cluster = clientes_clusters[clientes_clusters['cluster'] == cluster]['id'].tolist()
            clusters[cluster] = ids_cluster


        resultados_json = json.dumps(clusters)  
        resultados_json = json.loads(resultados_json)

        print(clusters)

        # os.remove(f"temp/{file.filename}")
            
        return resultados_json
